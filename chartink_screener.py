"""
╔══════════════════════════════════════════════════════════════╗
║   CHARTINK SCREENER AUTOMATION                               ║
║   Weekly 10/30 EMA breakout pair → saves to Excel             ║
╠══════════════════════════════════════════════════════════════╣
║  INSTALL ONCE (run in Command Prompt):                       ║
║   pip install requests beautifulsoup4 openpyxl pandas        ║
╚══════════════════════════════════════════════════════════════╝
"""

import sys, os, time, datetime

missing = []
for pkg in ['requests', 'bs4', 'openpyxl', 'pandas']:
    try: __import__(pkg)
    except ImportError: missing.append(pkg)
if missing:
    print(f"\n❌  Missing: {', '.join(missing)}")
    print(f"    Run:  pip install {' '.join(missing)}")
    if not os.environ.get("GITHUB_ACTIONS"):
        input("\nPress Enter to exit...")
    sys.exit(1)

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
#  ✏️  YOUR SCREENERS
#  Format: ("Display Name", "clause", "your chartink scan clause here")
# ══════════════════════════════════════════════════════════════════════════════
SCREENERS = [
    # Built to feed the weekly 10/30-EMA vision pipeline (fetchimages_nse.py /
    # vision_engine.py). "Ready Now" mirrors the vision prompt's Readiness =
    # Ready Now bucket: price above a rising 10-week EMA which itself sits
    # above a rising 30-week EMA, tagging a fresh ~10-week high, on volume
    # expansion, and not yet extended too far to chase. "Basing" is a looser
    # pre-trigger watchlist feed for stocks approaching but not yet through
    # that same ceiling - useful for catching setups a few days before they
    # actually trigger. Paste-test both in Chartink's scanner console before
    # relying on them if you change the thresholds - the "N period ago
    # <indicator>(...)" slope-check syntax hasn't been independently
    # verified against the live Chartink engine.
    (
        "Weekly 10-30 EMA Breakout - Ready Now",
        "clause",
        "( {cash} (  weekly close >  weekly ema( weekly close , 10 ) and  weekly ema( weekly close , 10 ) >  weekly ema( weekly close , 30 ) and  weekly ema( weekly close , 30 ) >  1 week ago weekly ema( weekly close , 30 ) and  daily high >=  weekly max( 10 , weekly high ) *  0.98 and  daily close <=  weekly ema( weekly close , 10 ) *  1.15 and  daily volume >  daily sma( daily volume , 20 ) *  1.3 and  daily close *  daily sma( daily volume , 20 ) >  5000000 and  market cap >  300 and  daily close >  10 ) )"
    ),
    (
        "Weekly 10-30 EMA Basing - Forming",
        "clause",
        "( {cash} (  weekly close >  weekly ema( weekly close , 10 ) and  weekly ema( weekly close , 10 ) >  weekly ema( weekly close , 30 ) *  0.98 and  weekly close >=  weekly max( 10 , weekly high ) *  0.85 and  weekly close <  weekly max( 10 , weekly high ) *  0.99 and  daily close *  daily sma( daily volume , 20 ) >  5000000 and  market cap >  300 and  daily close >  10 ) )"
    ),
]

PAUSE_BETWEEN = 5   # seconds between requests (keep ≥ 4 to avoid 419)

OUTPUT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    f"Chartink_Screener_{datetime.date.today().strftime('%Y-%m-%d')}.xlsx"
)

PROCESS_URL = "https://chartink.com/screener/process"

# ══════════════════════════════════════════════════════════════════════════════
#  🚫  INDEX & ETF EXCLUSION LIST
#  Tickers matching these exact names OR containing these substrings are dropped
# ══════════════════════════════════════════════════════════════════════════════
_EXCLUDE_EXACT = {
    # ── NSE Indices ────────────────────────────────────────────────────────────
    "NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "NIFTYNXT50",
    "NIFTY50", "NIFTY100", "NIFTY200", "NIFTY500", "NIFTYMIDCAP50",
    "NIFTYMIDCAP100", "NIFTYMIDCAP150", "NIFTYMIDCAP400",
    "NIFTYSMALLCAP50", "NIFTYSMALLCAP100", "NIFTYSMALLCAP250",
    "NIFTYMICROCAP250", "NIFTYLARGECAP", "SENSEX", "BSE500",
    "NIFTYAUTO", "NIFTYBANK", "NIFTYFIN", "NIFTYFINSERVICE",
    "NIFTYFMCG", "NIFTYIT", "NIFTYMEDIA", "NIFTYMETAL",
    "NIFTYPHARMA", "NIFTYREALTY", "NIFTYPSUBANK", "NIFTYINFRA",
    "NIFTYCPSE", "NIFTYMHC", "NIFTYENERGY", "NIFTYDIVOPPS50",
    "NIFTYALPHA50", "NIFTYQUALITY30", "NIFTYLOWVOL30",
    "NIFTY100LOWVOL30", "NIFTY50DIVPNTS", "NIFTYINDIALVOL",
    "INDIA VIX", "INDIAVIX",
    # ── Common ETFs on NSE ────────────────────────────────────────────────────
    "NIFTYBEES", "BANKBEES", "JUNIORBEES", "PSUBNKBEES", "ITBEES",
    "GOLDBEES", "SILVERBEES", "LIQUIDBEES", "LIQUIDCASE", "LIQUIDETF",
    "SETFNIF50", "SETFNN50", "SETFNIFIT", "SETFBSE100", "SETFNIFBK",
    "ICICIB22", "ICICITECH", "MOM30IETF", "MAFANG", "MOMOMENTUM",
    "MOSMALL250", "MONIFTY500", "MOVALUE", "MOQUALITY", "MOHEALTH",
    "NIFTIETF", "HDFCNIFTY", "HDFCMID150", "HDFCSML250", "HDFCSMALL",
    "HDFCLIQUID", "HDFCGOLD", "HDFCSILVER", "HDFCNIFIT", "HDFCNIFBAN",
    "SBIETFQLTY", "SBIETFPB", "SBIETFIT", "SBIETFCONS", "SBIETFPHARMA",
    "SBIETFMID150", "SBIETF1000",
    "KOTAKNIFTY", "KOTAKBANK", "KOTAKGOLD", "KOTAKSILVER",
    "KOTAKNIFBK", "KOTAKMID50", "KOTAKPSUBK", "KOTAKBANKSO",
    "AXISNIFTY", "AXISSMALL", "AXISNIFTYBK",
    "UTINIFTETF", "UTISENSETF", "UTIBANKETF", "UTINEXT50",
    "ABSLLIQUID", "ABSLPSE",
    "MIRAERASETF", "MIRAEASETBK", "MIRAEASETGD",
    "BSLGOLDETF", "BSLNIFTY",
    "TATANIFTY50", "TATADIGITAL", "TATASMCAP",
    "ITETF", "AUTIETF", "PHARMABEES", "INFRABEES",
    "CPSEETF", "CONSUMBEES", "PVTBANKETF",
    "NIFTYIETF", "NIF100IETF", "NIF100BEES", "NIFTYMID",
    "LOWVOLIETF", "ALPHAETF", "QUAL30IETF",
    "PSUBANK", "SHARIABEES",
}

_EXCLUDE_SUBSTRINGS = (
    "BEES", "ETF", "LIQUIDCASE", "LIQETF", "INDEX", "NIFTYBK",
    "IETF", "SETF",
)

def is_index_or_etf(ticker: str) -> bool:
    """Return True if the ticker looks like an index or ETF (name-based check)."""
    t = ticker.strip().upper()
    if t in _EXCLUDE_EXACT:
        return True
    for sub in _EXCLUDE_SUBSTRINGS:
        if sub in t:
            return True
    return False

def filter_stocks_only(df):
    """
    Drop rows that are indices or ETFs using two layers:
      1. Market Cap == 0 or null  ->  not a real stock (primary, most reliable)
      2. Name-pattern check       ->  catches anything with mcap missing from API
    Returns a clean df with only real equity stocks.
    """
    if df.empty:
        return df

    ticker_col = "Ticker" if "Ticker" in df.columns else df.columns[0]

    # Layer 1: market cap filter (ETFs/indices have 0 or null mcap in Chartink)
    if "Market Cap" in df.columns:
        def _mcap_ok(val):
            try:
                return float(str(val).replace(",", "").strip() or "0") > 0
            except Exception:
                return True   # if unparseable, keep and let layer 2 decide
        mcap_mask = df["Market Cap"].apply(_mcap_ok)
    else:
        mcap_mask = pd.Series([True] * len(df), index=df.index)

    # Layer 2: name pattern filter
    name_mask = ~df[ticker_col].astype(str).str.strip().apply(is_index_or_etf)

    before = len(df)
    df = df[mcap_mask & name_mask].reset_index(drop=True)
    removed = before - len(df)
    if removed:
        print(f"   🚫  Removed {removed} index/ETF row(s) — {len(df)} stocks remain")
    return df

# ══════════════════════════════════════════════════════════════════════════════
#  FETCH DATA FROM CHARTINK
# ══════════════════════════════════════════════════════════════════════════════
def fetch_chartink(session, scan_clause):
    try:
        page_headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }
        page = session.get(
            "https://chartink.com/screener/",
            headers=page_headers,
            timeout=20,
        )
        time.sleep(3)

        if page.status_code != 200:
            print(f"   ❌  Could not load chartink.com (HTTP {page.status_code})")
            return pd.DataFrame()

        soup = BeautifulSoup(page.text, "html.parser")
        meta = soup.find("meta", {"name": "csrf-token"})

        if meta and meta.get("content"):
            csrf_token = meta["content"]
            print(f"   🔑  CSRF from HTML meta tag: {csrf_token[:16]}...")
        else:
            raw = session.cookies.get("XSRF-TOKEN", "")
            csrf_token = requests.utils.unquote(raw)
            if csrf_token:
                print(f"   🔑  CSRF from cookie fallback: {csrf_token[:16]}...")
            else:
                print("   ❌  Could not find CSRF token — Chartink may have changed its page structure")
                return pd.DataFrame()

        post_headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            "Referer":          "https://chartink.com/screener/",
            "Origin":           "https://chartink.com",
            "X-Requested-With": "XMLHttpRequest",
            "X-CSRF-TOKEN":     csrf_token,
            "Content-Type":     "application/x-www-form-urlencoded; charset=UTF-8",
            "Accept":           "application/json, text/javascript, */*; q=0.01",
            "Accept-Language":  "en-US,en;q=0.9",
        }

        resp = session.post(
            PROCESS_URL,
            data={"scan_clause": scan_clause, "_token": csrf_token},
            headers=post_headers,
            timeout=25,
        )

        if resp.status_code == 419:
            print("   ❌  HTTP 419 — CSRF token rejected by Chartink.")
            print("       Possible fixes:")
            print("       1. Increase PAUSE_BETWEEN at the top of this script (try 8+)")
            print("       2. Open chartink.com in your browser and log in, then run again")
            print("       3. Wait a few minutes and try again")
            return pd.DataFrame()

        if resp.status_code != 200:
            print(f"   ❌  HTTP {resp.status_code}")
            return pd.DataFrame()

        try:
            data = resp.json()
        except Exception:
            if "<html" in resp.text.lower() or "<!doctype" in resp.text.lower():
                print("   ❌  Chartink returned an HTML page — you may need to log in via browser")
                print(f"       Response snippet: {resp.text[:200]}")
            else:
                print(f"   ❌  Could not parse response. Snippet: {resp.text[:300]}")
            return pd.DataFrame()

        if "data" not in data:
            print(f"   ⚠️  Unexpected response keys: {list(data.keys())}")
            print(f"       Full response: {str(data)[:300]}")
            return pd.DataFrame()

        raw_data = data["data"]
        print(f"   📊  Records received: {len(raw_data)}")
        if not raw_data:
            print("   ⚠️  API returned 0 rows")
            return pd.DataFrame()

        # ── ONE-TIME DEBUG: print all raw fields from Chartink ────────────────
        if not getattr(fetch_chartink, "_fields_printed", False):
            fetch_chartink._fields_printed = True
            sample = raw_data[0]
            print(f"\n   🔍  RAW CHARTINK FIELDS (all keys in API response):")
            for k, v in sample.items():
                print(f"       {k!r:30s} = {str(v)[:60]!r}")
            print()

        df = pd.DataFrame(raw_data)

        rename_map = {
            "nsecode":  "Ticker",
            "bsecode":  "BSE Code",
            "per_chg":  "% Change",
            "close":    "Close",
            "volume":   "Volume",
            "turnover": "Turnover (Cr)",
            "mcap":     "Market Cap",
            "sr_no":    "Sr No",
        }
        df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

        if "Ticker" in df.columns:
            cols = ["Ticker"] + [c for c in df.columns if c != "Ticker"]
            df   = df[cols]

        return df

    except Exception as e:
        print(f"   ❌  Unexpected error: {e}")
        import traceback; traceback.print_exc()
        return pd.DataFrame()

# ══════════════════════════════════════════════════════════════════════════════
#  BUILD EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(screener_results, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    def fill(h):
        return PatternFill("solid", fgColor=h)

    def thin(color="CCCCCC"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def medium(color="666666"):
        s = Side(style="medium", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    C_DARK  = "1A3A5C"; C_HDR   = "2E5FA3"; C_GOLD  = "FFD700"; C_WHITE = "FFFFFF"
    C_S1    = "FFFFFF"; C_S2    = "EBF3FB"; C_TICK  = "DDEEFF"
    C_POSB  = "E6F4EA"; C_POSF  = "137333"
    C_NEGB  = "FCE8E6"; C_NEGF  = "C5221F"
    C_NOTE  = "FFFBEA"; C_GREEN = "00763D"; C_RED   = "C00000"

    now_str = datetime.datetime.now().strftime("%d %b %Y  %H:%M")
    all_tickers_flat = []
    tab_colors = ["2E75B6","70AD47","C00000","FF8C00","7030A0",
                  "00B0F0","BF9000","375623","4472C4","ED7D31",
                  "A9D18E","0070C0","7F7F7F","FFC000","FF0000"]

    for idx, (name, df) in enumerate(screener_results):
        safe = (name[:31].replace("/","_").replace("\\","_")
                .replace("?","").replace("*","")
                .replace("[","").replace("]","").replace(":",""))
        ws = wb.create_sheet(title=safe)
        ws.sheet_properties.tabColor = tab_colors[idx % len(tab_colors)]

        n_cols   = max(len(df.columns), 4) if not df.empty else 4
        last_col = get_column_letter(n_cols)

        ws.row_dimensions[1].height = 5
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 14

        ws.merge_cells(f"A2:{last_col}2")
        c = ws["A2"]
        c.value     = f"  {name.upper()}   |   {len(df)} stocks   |   {now_str}"
        c.font      = Font(name="Arial", bold=True, size=12, color=C_WHITE)
        c.fill      = fill(C_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(f"A3:{last_col}3")
        c = ws["A3"]
        c.value     = f"  Source: chartink.com   |   Run: {now_str}"
        c.font      = Font(name="Arial", italic=True, size=8, color="AAAAAA")
        c.fill      = fill(C_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")

        if df.empty:
            ws["A5"].value = "⚠️  No data retrieved for this screener"
            ws["A5"].font  = Font(name="Arial", size=11, color=C_RED)
            continue

        # ── Drop indices / ETFs — keep real stocks only ───────────────────────
        df = filter_stocks_only(df)
        ws["A2"].value = f"  {name.upper()}   |   {len(df)} stocks   |   {now_str}"
        if df.empty:
            ws["A5"].value = "⚠️  No stocks after excluding indices/ETFs"
            ws["A5"].font  = Font(name="Arial", size=11, color=C_RED)
            continue

        pct_cols = {col for col in df.columns
                    if col == "% Change"
                    or df[col].dropna().astype(str).head(20).str.contains(r"%").mean() > 0.3}

        num_cols = set()
        for col in df.columns:
            if col in pct_cols: continue
            try:
                pd.to_numeric(df[col].dropna().head(20), errors="raise")
                num_cols.add(col)
            except Exception:
                pass

        ticker_col  = "Ticker" if "Ticker" in df.columns else df.columns[0]
        ticker_cidx = list(df.columns).index(ticker_col) + 1

        ws.row_dimensions[4].height = 22
        for ci, col_name in enumerate(df.columns, 1):
            c = ws.cell(row=4, column=ci)
            c.value     = col_name
            c.font      = Font(name="Arial", bold=True, size=9, color=C_WHITE)
            c.fill      = fill(C_HDR)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = medium()
            ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col_name)) + 3, 10)

        for ri, (_, row) in enumerate(df.iterrows()):
            er = ri + 5
            ws.row_dimensions[er].height = 15
            bg = C_S1 if ri % 2 == 0 else C_S2

            for ci, (col_name, val) in enumerate(zip(df.columns, row), 1):
                c   = ws.cell(row=er, column=ci)
                txt = str(val).strip() if pd.notna(val) else ""
                c.value = txt

                if ci == ticker_cidx:
                    c.font      = Font(name="Arial", bold=True, size=9, color=C_DARK)
                    c.fill      = fill(C_TICK)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border    = thin("AACCEE")
                elif col_name in pct_cols and txt:
                    try:
                        neg = float(txt.replace("%","").replace(",","")) < 0
                    except Exception:
                        neg = txt.startswith("-")
                    c.font      = Font(name="Arial", size=9, color=C_NEGF if neg else C_POSF)
                    c.fill      = fill(C_NEGB if neg else C_POSB)
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.border    = thin()
                elif col_name in num_cols and txt:
                    c.font      = Font(name="Arial", size=9)
                    c.fill      = fill(bg)
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.border    = thin()
                else:
                    c.font      = Font(name="Arial", size=9)
                    c.fill      = fill(bg)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border    = thin()

                cur_w = ws.column_dimensions[get_column_letter(ci)].width
                ws.column_dimensions[get_column_letter(ci)].width = min(
                    max(cur_w, len(txt) + 2), 30
                )

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{last_col}{4 + len(df)}"

        tickers = df[ticker_col].dropna().astype(str).str.strip().tolist()
        tickers = [t for t in tickers if t not in ("", "nan", "NaN")]
        tickers = [t for t in tickers if not is_index_or_etf(t)]
        all_tickers_flat.extend(tickers)

    # ── Summary tab ───────────────────────────────────────────────────────────
    from collections import Counter
    counts = Counter(t.upper() for t in all_tickers_flat)
    seen, unique = set(), []
    for t in all_tickers_flat:
        u = t.upper()
        if u not in seen and not is_index_or_etf(u):
            seen.add(u); unique.append(u)
    unique.sort()

    ws_s = wb.create_sheet(title="All Tickers")
    ws_s.sheet_properties.tabColor = "FFD700"
    ws_s.row_dimensions[1].height  = 5
    ws_s.row_dimensions[2].height  = 30
    ws_s.row_dimensions[3].height  = 14

    ws_s.merge_cells("A2:G2")
    c = ws_s["A2"]
    c.value     = f"  ALL TICKERS — DEDUPLICATED   |   {len(unique)} unique   |   {now_str}"
    c.font      = Font(name="Arial", bold=True, size=13, color="000000")
    c.fill      = fill(C_GOLD)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws_s.merge_cells("A3:G3")
    c = ws_s["A3"]
    c.value     = f"  Total raw: {len(all_tickers_flat)}   |   After dedup: {len(unique)}"
    c.font      = Font(name="Arial", italic=True, size=9, color="555555")
    c.fill      = fill(C_GOLD)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws_s.row_dimensions[5].height = 18
    ws_s.merge_cells("A5:G5")
    c = ws_s["A5"]
    c.value     = "  ✂  COPY-PASTE READY — Comma-Separated Ticker List"
    c.font      = Font(name="Arial", bold=True, size=10, color=C_WHITE)
    c.fill      = fill(C_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws_s.row_dimensions[6].height = 80
    ws_s.merge_cells("A6:G6")
    c = ws_s["A6"]
    c.value     = ", ".join(unique)
    c.font      = Font(name="Courier New", size=9)
    c.fill      = fill(C_NOTE)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    bd = Side(style="medium", color=C_HDR)
    c.border    = Border(left=bd, right=bd, top=bd, bottom=bd)

    for cl, w in [("A",6),("B",14),("C",20),("D",12),("E",12),("F",12),("G",12)]:
        ws_s.column_dimensions[cl].width = w

    ws_s.row_dimensions[8].height = 20
    for cl, hdr in [("A","#"),("B","Ticker"),("C","In # Screeners"),("D",""),("E",""),("F",""),("G","")]:
        c = ws_s[f"{cl}8"]
        c.value     = hdr
        c.font      = Font(name="Arial", bold=True, size=9, color=C_WHITE)
        c.fill      = fill(C_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = medium()

    for i, ticker in enumerate(unique):
        r   = 9 + i
        bg  = C_S1 if i % 2 == 0 else C_S2
        cnt = counts.get(ticker, 1)
        ws_s.row_dimensions[r].height = 15
        for cl, val in [("A", i + 1), ("B", ticker), ("C", cnt)]:
            c = ws_s[f"{cl}{r}"]
            c.value     = val
            c.font      = Font(name="Arial", size=9, bold=(cl == "B"),
                               color=C_GREEN if cnt > 1 else "000000")
            c.fill      = fill(C_S2 if cnt > 1 else bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin()

    ws_s.freeze_panes = "A9"
    wb.move_sheet("All Tickers", offset=-len(wb.sheetnames) + 1)

    wb.save(output_path)
    print(f"\n✅  Excel saved: {output_path}")
    return unique

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("  CHARTINK SCREENER AUTOMATION")
    print(f"  Total screeners: {len(SCREENERS)}")
    print("=" * 60)

    session = requests.Session()
    all_results = []

    for name, mode, value in SCREENERS:
        print(f"\n📡  Running: {name}")
        if mode.lower() != "clause":
            print("   ⚠️  Only 'clause' mode supported — skipping")
            continue

        df = fetch_chartink(session, value)
        if df.empty:
            print("   ⚠️  No results")
            all_results.append((name, df))
        else:
            print(f"   ✅  {len(df)} stocks found")
            all_results.append((name, df))

        time.sleep(PAUSE_BETWEEN)

    if not all_results:
        print("\n❌  No data retrieved.")
        if not os.environ.get("GITHUB_ACTIONS"):
            input("\nPress Enter to exit...")
        sys.exit(1)

    print(f"\n\n📊  Building Excel...")
    unique = build_excel(all_results, OUTPUT_FILE)

    print(f"\n🎯  {len(unique)} unique tickers")
    print(f"    Preview: {', '.join(unique[:12])}{'...' if len(unique) > 12 else ''}")

    if not os.environ.get("GITHUB_ACTIONS"):
        try:
            os.startfile(OUTPUT_FILE)
        except Exception:
            import subprocess
            subprocess.Popen(["start", OUTPUT_FILE], shell=True)
        input("\nPress Enter to close...")
    else:
        print(f"::notice::Excel file ready at {OUTPUT_FILE}")
